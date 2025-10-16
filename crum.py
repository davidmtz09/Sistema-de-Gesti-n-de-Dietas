import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from tkcalendar import Calendar
from datetime import datetime
import menu_principal
from utilerias import get_file_path
from pathlib import Path

def regresar():
    window.destroy()
    menu_principal.mostrar_menu()

def crum():
    global window

    class PeriodPopup:
        def __init__(self, parent, entry):
            self.top = tk.Toplevel(parent)
            self.top.title("Selecciona el periodo de alimento")
            self.top.geometry("400x550")
            self.top.transient(parent)
            self.top.grab_set()

            self.entry = entry

            tk.Label(self.top, text="Seleccione fecha de inicio", font=("Helvetica", 10)).pack(pady=10)
            today = datetime.today()
            self.calendar1 = Calendar(self.top, selectmode='day', year=today.year, month=today.month, day=today.day, locale='es_ES')
            self.calendar1.pack(pady=10, padx=20)

            tk.Label(self.top, text="Seleccione fecha de termino", font=("Helvetica", 10)).pack(pady=10)
            self.calendar2 = Calendar(self.top, selectmode='day', year=today.year, month=today.month, day=today.day, locale='es_ES')
            self.calendar2.pack(pady=10, padx=20)

            select_button = tk.Button(self.top, text="Seleccionar", command=self.select_dates)
            select_button.pack(pady=10)

        def select_dates(self):
            date1 = self.calendar1.get_date()
            date2 = self.calendar2.get_date()

            try:
                formatted_date1 = datetime.strptime(date1, "%d/%m/%y").strftime("%d/%m/%Y")
                formatted_date2 = datetime.strptime(date2, "%d/%m/%y").strftime("%d/%m/%Y")
            except ValueError:
                formatted_date1 = datetime.strptime(date1, "%d/%m/%y").strftime("%d/%m/%Y")
                formatted_date2 = datetime.strptime(date2, "%d/%m/%y").strftime("%d/%m/%Y")

            self.entry.delete(0, tk.END)
            self.entry.insert(0, f"{formatted_date1} - {formatted_date2}")
            self.top.destroy()

    def show_period_popup(entry):
        PeriodPopup(window, entry)

    def load_data():
        global tree, cols, template_data
        path = Path.home() / "ContPerFactura" / "crum.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active

        list_values = list(sheet.values)
        cols = list_values[0]

        # Cambiar el nombre del último encabezado a "Periodo_alimento"
        cols = list(cols)
        cols[-1] = "Periodo_alimento"
        cols[-3] = "SERVICIO"

        #Agregar estilos a treeview
        style = ttk.Style()
        #Tema
        style.theme_use("clam")
        #Configuracion de los colores del treeview
        style.configure("Treeview",
            background="#b7eefa",
            foreground="black")

        style.map("Treeview",
            background=[('selected', '#bd1323')])

        # Frame for the Treeview and Scrollbar
        tree_frame = tk.Frame(window)
        tree_frame.pack(expand=True, fill='both')

        # Create Treeview
        tree = ttk.Treeview(tree_frame, style="Treeview", columns=cols, show="headings")

        # Scrollbar
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        vsb.pack(side='right', fill='y')
        tree.configure(yscrollcommand=vsb.set)

        # Adjust column widths
        column_widths = {
            0: 10,    # Column 0
            1: 240,   # Column 1
            2: 10,    # Column 2
            3: 150,   # Column 3
            4: 10,    # Column 4
            5: 60,    # Column 5
            7: 70,   # Column 7
            8: 70,    # Column 8
            9: 40,    # Column 9
            10: 100,  # Column 10
        }
        
        for i, col_name in enumerate(cols):
            if i in column_widths:
                tree.column(col_name, width=column_widths[i], anchor='center')
            else:
                tree.column(col_name, width=100, anchor='center')      
            tree.heading(col_name, text=col_name)

        tree.pack(expand=True, fill='both')
        
        for value_tuple in list(sheet.values)[1:]:
            tree.insert('', tk.END, values=value_tuple)
        
        template_path = Path.home() / "ContPerFactura" / "plantilla.xlsx"
        template_workbook = openpyxl.load_workbook(template_path)
        template_sheet = template_workbook.active
        template_values = list(template_sheet.values)
        template_cols = template_values[0]
        template_data = {str(row[0]): row[1:] for row in template_values[1:]}

    def add_data():
        no_tarj = entries[0].get().strip()  # NO. TARJ
        new_period = entries[-1].get().strip()  # Periodo_alimento

        # Verificar que los campos obligatorios estén llenos
        if not no_tarj or not new_period:
            messagebox.showwarning("Advertencia", "Los campos 'NO. TARJ' y 'Periodo_alimento' son obligatorios.")
            return

        # Convertir las fechas del nuevo periodo
        try:
            new_start, new_end = map(lambda d: datetime.strptime(d.strip(), "%d/%m/%Y"), new_period.split("-"))
        except ValueError:
            messagebox.showerror("Error", "El formato del periodo es incorrecto. Use el formato 'DD/MM/YYYY - DD/MM/YYYY'.")
            return

        # Validar solapamientos
        for row_id in tree.get_children():
            row_values = tree.item(row_id)["values"]
            existing_tarj = str(row_values[0]).strip()
            existing_period = str(row_values[-1]).strip()

            if existing_tarj == no_tarj:
                try:
                    existing_start, existing_end = map(lambda d: datetime.strptime(d.strip(), "%d/%m/%Y"), existing_period.split("-"))
                    # Comprobar si los periodos se solapan
                    if new_start <= existing_end and new_end >= existing_start:
                        messagebox.showwarning("Advertencia", f"El periodo ingresado se repite con un registro existente para el NO. TARJ: {no_tarj}.")
                        return
                except ValueError:
                    continue

        # Si pasa las validaciones, agregar el nuevo registro
        new_data = [entry.get().upper() for entry in entries]
        tree.insert('', tk.END, values=new_data)
        for entry in entries:
            entry.delete(0, tk.END)

    def delete_data():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Advertencia", "Seleccione un registro para eliminar")
            return

        tree.delete(selected_item)

    def save_data():
        path = Path.home() / "ContPerFactura" / "crum.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        sheet.delete_rows(2, sheet.max_row)  # Eliminar datos existentes, excepto el encabezado

        for row in tree.get_children():
            sheet.append(tree.item(row)['values'])
        
        workbook.save(path)
        messagebox.showinfo("Información", "Cambios guardados correctamente")

    def autocomplete(event):
        input_value_1 = entries[0].get().strip()
        input_value_2 = entries[1].get().strip()
        
        if input_value_1:
            input_value_1 = str(input_value_1) 
            if input_value_1 in template_data:
                row_data = template_data[input_value_1]
                for i, value in enumerate(row_data):
                    if i < len(entries) - 3:  
                        entries[i + 1].delete(0, tk.END)
                        if value is not None:
                            entries[i + 1].insert(0, value)
                        else:
                            entries[i + 1].delete(0, tk.END)

        if input_value_2:
            for key, values in template_data.items():
                if values[1] == input_value_2:
                    for i, value in enumerate([key] + values[1:]):
                        if i < len(entries) - 3: 
                            entries[i].delete(0, tk.END)
                            if value is not None:
                                entries[i].insert(0, value)
                            else:
                                entries[i].delete(0, tk.END)
                    break

    window = tk.Tk()
    window.title("REGISTRO DEL PERSONAL CRUM")
    window.geometry("1500x600")
    load_data()

    # Frame
    input_frame = tk.Frame(window)
    input_frame.pack(pady=10, fill='x')

    cols = tree['column']
    entries = []
    for i, col in enumerate(cols):
        label = tk.Label(input_frame, text=col if col != cols[-1] else "Periodo_alimento")
        label.grid(row=0, column=i, padx=5, pady=5)
        if i == len(cols) - 1:  # Last column for date range
            period_entry = tk.Entry(input_frame)
            period_entry.bind("<Button-1>", lambda e, ent=period_entry: show_period_popup(ent))
            entry = period_entry
        elif i == len(cols) - 2:  # Penultimate column for combobox DESAYUNO, COMIDA, CENA
            entry = ttk.Combobox(input_frame, values=["COMIDA", "CENA"], width=10, state="readonly", name="alimento")
        elif i == len(cols) - 3:  # Antepenultimate column for combobox 
            entry = ttk.Combobox(input_frame, values=["CRUM"], width=10, state="readonly", name="tipo")
            entry.set("CRUM")  # Establece el valor predeterminado
        else:
            entry = tk.Entry(input_frame)
        entry.grid(row=1, column=i, padx=5, pady=5)
        entries.append(entry)

    entries[0].bind("<FocusOut>", autocomplete)
    entries[1].bind("<FocusOut>", autocomplete)

    # Frame for buttons
    button_frame = tk.Frame(window)
    button_frame.pack(pady=10)

    # Add data button
    add_button = tk.Button(button_frame, text="Insertar Registro", command=add_data, bg="purple", fg="white")
    add_button.pack(side=tk.LEFT, padx=5)

    # Delete data button
    delete_button = tk.Button(button_frame, text="Eliminar Registro", command=delete_data, bg="purple", fg="white")
    delete_button.pack(side=tk.LEFT, padx=5)

    # Save data button
    save_button = tk.Button(button_frame, text="Guardar Cambios", command=save_data, bg="purple", fg="white")
    save_button.pack(side=tk.LEFT, padx=5)

    regresar_button = tk.Button(button_frame, text="Regresar al Menú", command=regresar, bg="purple", fg="white")
    regresar_button.pack(side=tk.LEFT, padx=5)

    window.mainloop()