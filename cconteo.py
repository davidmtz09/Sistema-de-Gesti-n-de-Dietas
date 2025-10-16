import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from calendar import monthrange
import pandas as pd
import webbrowser
import tempfile
import os
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import menu_principal
from utilerias import get_file_path
from pathlib import Path

def nombre_a_numero_mes(nombre_mes):
    nombres_meses = {
        "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4, "Mayo": 5, "Junio": 6,
        "Julio": 7, "Agosto": 8, "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12
    }
    return nombres_meses.get(nombre_mes, 0)  # Retorna 0 si el mes no es válido

def regresar():
    root.destroy()
    menu_principal.mostrar_menu()

def conteo():
    global root
    # Meses
    nombres_meses = [
        "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]

    def obtener_nombre_mes(numero_mes):
        """Devuelve el nombre del mes en español dado su número (1-12)."""
        return nombres_meses[numero_mes]

    # Función para contar alimentos por día
    def contar_alimentos_por_dia(df, filtro_columna_8, dia, mes, anho):
        conteos = {"DESAYUNO": 0, "COMIDA": 0, "CENA": 0}
        # Filtrar la columna 8 si es necesario (aplica para Internos, Residentes y otras categorías específicas)
        if filtro_columna_8:
            df_filtrado = df[df.iloc[:, 8] == filtro_columna_8]
        else:
            df_filtrado = df  # Para vales y suplencias, no se filtra por columna 8

        for _, row in df_filtrado.iterrows():
            periodo = row.iloc[10]
            tipo_alimento = row.iloc[9]

            try:
                fecha_inicio, fecha_fin = periodo.split(' - ')
                fecha_inicio = pd.to_datetime(fecha_inicio, dayfirst=True)
                fecha_fin = pd.to_datetime(fecha_fin, dayfirst=True)

                # Verificar si el día está dentro del rango de fechas
                if fecha_inicio <= pd.Timestamp(anho, mes, dia) <= fecha_fin:
                    conteos[tipo_alimento] += 1

            except ValueError:
                continue

        return conteos

    # Función para actualizar las tablas
    def update_tables():
        try:
            year = int(year_var.get())
            month = nombre_a_numero_mes(month_var.get())
            num_days = monthrange(year, month)[1]

            for widget in tables_canvas_frame.winfo_children():
                widget.destroy()

            # Define la ruta a la carpeta ContPerFactura
            user_folder = Path.home() / "ContPerFactura"

            # Cargar los DataFrames desde los archivos Excel en la carpeta ContPerFactura
            pbase_df = pd.read_excel(user_folder / "pbase.xlsx")
            pcontrato_df = pd.read_excel(user_folder / "pcontrato.xlsx")
            medicos_df = pd.read_excel(user_folder / "medicos.xlsx")
            suplencias_df = pd.read_excel(user_folder / "suplencias.xlsx")
            vales_df = pd.read_excel(user_folder / "vales.xlsx")
            hrf_df = pd.read_excel(user_folder / "phrf.xlsx")
            imssb_df = pd.read_excel(user_folder / "imssb.xlsx")
            crum_df = pd.read_excel(user_folder / "crum.xlsx")
            pasantes_df = pd.read_excel(user_folder / "pasantes.xlsx")

            # Procesar cada día del mes seleccionado
            for day in range(1, num_days + 1):
                date_str = f"{day:02d}/{month:02d}/{year}"
                day_frame = tk.Frame(tables_canvas_frame, borderwidth=2, relief="solid", padx=10, pady=10)
                day_frame.grid(row=(day-1)//3, column=(day-1)%3, padx=5, pady=5, sticky="nsew")

                # Crear la tabla para el día
                columns = ["Fecha", "Desayuno", "Comida", "Cena"]
                table = ttk.Treeview(day_frame, columns=columns, show="headings", height=24)

                # Configurar encabezados de la tabla
                table.heading("Fecha", text=date_str)
                table.column("Fecha", width=150, anchor="center")

                for col in columns[1:]:
                    table.heading(col, text=col)
                    table.column(col, width=100, anchor="center")

                # Realizar los conteos para cada categoría
                categorias = {
                    "GAB": contar_alimentos_por_dia(pbase_df, "GAB", day, month, year),
                    "GAC": contar_alimentos_por_dia(pcontrato_df, "GAC", day, month, year),
                    "GA-HRF": contar_alimentos_por_dia(hrf_df, "GA-HRF", day, month, year),
                    "GA-IMSSB": contar_alimentos_por_dia(imssb_df, "GA-IMSSB", day, month, year),
                    "GBB": contar_alimentos_por_dia(pbase_df, "GBB", day, month, year),
                    "GBC": contar_alimentos_por_dia(pcontrato_df, "GBC", day, month, year),
                    "GB-HRF": contar_alimentos_por_dia(hrf_df, "GB-HRF", day, month, year),
                    "GB-IMSSB": contar_alimentos_por_dia(imssb_df, "GB-IMSSB", day, month, year),
                    "JADB": contar_alimentos_por_dia(pbase_df, "JADB", day, month, year),
                    "JADC": contar_alimentos_por_dia(pcontrato_df, "JADC", day, month, year),
                    "JAD-HRF": contar_alimentos_por_dia(hrf_df, "JAD-HRF", day, month, year),
                    "JAD-IMSSB": contar_alimentos_por_dia(imssb_df, "JAD-IMSSB", day, month, year),
                    "JAD-PASANTES": contar_alimentos_por_dia(pasantes_df, "JAD-PASANTES", day, month, year),
                    "JANB": contar_alimentos_por_dia(pbase_df, "JANB", day, month, year),
                    "JANC": contar_alimentos_por_dia(pcontrato_df, "JANC", day, month, year),
                    "JAN-HRF": contar_alimentos_por_dia(hrf_df, "JAN-HRF", day, month, year),
                    "JAN-IMSSB": contar_alimentos_por_dia(imssb_df, "JAN-IMSSB", day, month, year),
                    "CRUM": contar_alimentos_por_dia(crum_df, "CRUM", day, month, year),
                    "Internos": contar_alimentos_por_dia(medicos_df, "MIP", day, month, year),
                    "Residentes": contar_alimentos_por_dia(medicos_df, "MR", day, month, year),
                    "Suplencias": contar_alimentos_por_dia(suplencias_df, "", day, month, year),
                    "Vales": contar_alimentos_por_dia(vales_df, "", day, month, year),
                }

                subtotales = {"DESAYUNO": 0, "COMIDA": 0, "CENA": 0}

                # Insertar filas con los textos especificados
                for categoria, conteo in categorias.items():
                    row_values = [categoria, conteo["DESAYUNO"], conteo["COMIDA"], conteo["CENA"]]
                    table.insert('', 'end', values=row_values)
                    for key in subtotales:
                        subtotales[key] += conteo[key]

                # Añadir la fila de subtotales
                label_style1 = {"font": ("Arial", 10, "bold")}
                table.insert('', 'end', values=("Subtotales", subtotales["DESAYUNO"], subtotales["COMIDA"], subtotales["CENA"]), tags=('subtotales',))
                table.tag_configure('subtotales', background="#c965d3", foreground="white", **label_style1)

                # Añadir la fila de total consumos
                total_consumos = sum(subtotales.values())
                table.insert('', 'end', values=("Total consumos", "", total_consumos, ""), tags=('total_consumos',))
                table.pack(padx=5, pady=5, fill=tk.BOTH, expand=True)
                table.tag_configure('total_consumos', background="purple", foreground="white", **label_style1)

                # Ajustar el tamaño del marco del día para que la tabla ocupe todo el ancho disponible
                day_frame.update_idletasks()
                day_frame_width = table.winfo_width()
                day_frame.config(width=day_frame_width)

            # Actualizar el tamaño del canvas
            tables_canvas.configure(scrollregion=tables_canvas.bbox("all"))

        except ValueError:
            messagebox.showerror("Error", "Asegúrese de que el mes y el año estén seleccionados correctamente.")
        except FileNotFoundError as e:
            messagebox.showerror("Error", f"No se pudo encontrar el archivo: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"Ha ocurrido un error: {str(e)}")

    # Función para exportar a PDF
    def exportar_pdf():
        try:
            mes = nombre_a_numero_mes(month_var.get())
            anho = int(year_var.get())
            nombre_mes = obtener_nombre_mes(mes)

            # Nombre del archivo PDF
            file_name = get_file_path("reporte.pdf")
            pdf = SimpleDocTemplate(file_name, pagesize=letter)

            elements = []
            styles = getSampleStyleSheet()

            # Recorrer todas las tablas en la interfaz y agregar cada una a una nueva "hoja" con encabezado
            for widget in tables_canvas_frame.winfo_children():
                # Encabezado del hospital
                header_text = """
                HOSPITAL GENERAL DE CUERNAVACA DR. JOSE G. PARRES<br/>
                SUBDIRECCIÓN MÉDICA<br/>
                DEPARTAMENTO DE NUTRICIÓN
                """
                header_paragraph = Paragraph(f"<para align='center'>{header_text}</para>", styles['Normal'])
                elements.append(header_paragraph)
                elements.append(Paragraph("<br/><br/>", styles['Normal']))

                # Título
                title = Paragraph(f"Concentrado Mensual de Alimentos del Personal {nombre_mes} {anho}", styles['Title'])
                elements.append(title)
                elements.append(Paragraph("<br/>", styles['Normal']))

                # Crear datos de la tabla
                table_data = []

                # Encabezado de la tabla con la fecha
                date_str = widget.winfo_children()[0].heading("Fecha", "text")
                table_data.append([date_str, "Desayuno", "Comida", "Cena"])

                # Obtener los datos de la tabla en la interfaz
                tree = widget.winfo_children()[0]  # El Treeview es el primer widget en el frame
                for child in tree.get_children():
                    row = tree.item(child)['values']
                    table_data.append(row)

                # Crear tabla para el PDF
                table = Table(table_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch])

                # Agregar estilo a la tabla
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
                    ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ]))

                # Añadir la tabla a los elementos
                elements.append(table)
                elements.append(Paragraph("<br/><br/>", styles['Normal']))

                # Añadir salto de página para simular "hojas" separadas
                elements.append(PageBreak())

            # Generar el PDF
            pdf.build(elements)

            # Abrir el PDF generado
            webbrowser.open(file_name)

        except ValueError as ve:
            print(f"Error de valor: {ve}")  # Para depurar errores de conversión
            messagebox.showerror("Error", "Asegúrese de que el mes y el año estén seleccionados correctamente.")
        except FileNotFoundError as e:
            messagebox.showerror("Error", f"No se pudo encontrar el archivo: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"Ha ocurrido un error: {str(e)}")

    def exportar_excel():
        try:
            mes = nombre_a_numero_mes(month_var.get())
            anho = int(year_var.get())
            nombre_mes = obtener_nombre_mes(mes)

            # Crear un DataFrame para cada día del mes
            datos_excel = []
            for widget in tables_canvas_frame.winfo_children():
                # Obtener la fecha del encabezado de la tabla
                fecha = widget.winfo_children()[0].heading("Fecha", "text")

                # Recorrer las filas de la tabla
                tree = widget.winfo_children()[0]
                for child in tree.get_children():
                    row = tree.item(child)['values']
                    datos_excel.append([fecha] + row)

            # Crear DataFrame con los datos recopilados
            df = pd.DataFrame(datos_excel, columns=["Fecha", "Servicio", "Desayuno", "Comida", "Cena"])

            # Crear un archivo Excel temporal
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            df.to_excel(temp_file.name, index=False, sheet_name=f"{nombre_mes}_{anho}")

            # Cargar el archivo Excel generado para aplicar estilos
            wb = load_workbook(temp_file.name)
            ws = wb.active

            # Aplicar color morado a los encabezados y texto en blanco
            header_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # Morado
            header_font = Font(color="FFFFFF", bold=True)  # Blanco y negrita

            for cell in ws[1]:  # Primera fila son los encabezados
                cell.fill = header_fill
                cell.font = header_font

            # Hacer que el encabezado sea estático al hacer scroll
            ws.freeze_panes = "A2"  # Fija la primera fila

            # Aplicar un color distinto a las filas de "Total consumos"
            total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[1].value == "Total consumos":  # Suponiendo que "Categoría" es la segunda columna
                    for cell in row:
                        cell.fill = total_fill

            # Ajustar el ancho de las columnas
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter  # Obtener la letra de la columna
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[col_letter].width = adjusted_width

            # Guardar los cambios y cerrar el archivo
            wb.save(temp_file.name)
            wb.close()

            # Abrir el archivo Excel generado
            os.startfile(temp_file.name)

        except ValueError:
            messagebox.showerror("Error", "Asegúrese de que el mes y el año estén seleccionados correctamente.")
        except FileNotFoundError as e:
            messagebox.showerror("Error", f"No se pudo encontrar el archivo: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"Ha ocurrido un error: {str(e)}")

    # Crear ventana
    root = tk.Tk()
    root.title("CONTEO")
    root.geometry("1200x600")

    # Agregar estilos al Treeview
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Treeview",
        background="#b7eefa",
        foreground="black",
        fieldbackground="#b7eefa"
    )
    style.map("Treeview",
        background=[('selected', '#bd1323')],
        foreground=[('selected', 'white')]
    )

    # Variables de control
    month_var = tk.StringVar()
    year_var = tk.StringVar()

    # Frame para la selección de mes y año
    frame_select = tk.Frame(root)
    frame_select.pack(pady=10)

    # Etiqueta principal en el centro con fondo transparente (si el sistema lo permite)
    tk.Label(frame_select, text="HOSPITAL GENERAL DE CUERNAVACA DR. JOSE G. PARRES \n SUBDIRECCIÓN MÈDICA \n DEPARTAMENTO DE NUTRICIÓN", 
            font=('Arial', 10), anchor="center", justify="center", bg=frame_select.cget("background")).grid(row=0, column=0, columnspan=4, padx=5, pady=10, sticky="ew")

    # Etiqueta del título con fondo transparente (si el sistema lo permite)
    tk.Label(frame_select, text="CONCENTRADO MENSUAL DE ALIMENTOS DEL PERSONAL POR DIA", 
            font=('Verdana', 12, "bold"), anchor="center", justify="center", bg=frame_select.cget("background")).grid(row=1, column=0, columnspan=4, padx=5, pady=10, sticky="ew")

    # Etiqueta y menú desplegable para el mes
    tk.Label(frame_select, text="Seleccione Mes:").grid(row=2, column=1, padx=5, pady=5, sticky="e")
    month_menu = ttk.Combobox(frame_select, textvariable=month_var, values=[
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ], state="readonly")
    month_menu.grid(row=2, column=2, padx=5, pady=5, sticky="w")

    # Etiqueta y menú desplegable para el año
    tk.Label(frame_select, text="Seleccione Año:").grid(row=3, column=1, padx=5, pady=5, sticky="e")
    year_menu = ttk.Combobox(frame_select, textvariable=year_var, values=[str(year) for year in range(2024, 2051)], state="readonly")
    year_menu.grid(row=3, column=2, padx=5, pady=5, sticky="w")

    # Botón para actualizar las tablas
    update_button = tk.Button(frame_select, text="Actualizar", command=update_tables, bg="#6a0dad", fg="white")
    update_button.grid(row=4, column=0, padx=5, pady=10, sticky="ew")

    # Botón para exportar a PDF
    pdf_button = tk.Button(frame_select, text="Exportar a PDF", command=exportar_pdf, bg="#bd1323", fg="white")
    pdf_button.grid(row=4, column=1, padx=5, pady=10, sticky="ew")

    # Botón para exportar a Excel
    excel_button = tk.Button(frame_select, text="Exportar a EXCEL", command=exportar_excel, bg="green", fg="white")
    excel_button.grid(row=4, column=2, padx=5, pady=10, sticky="ew")

    # Botón para regresar al menú
    regresar_button = tk.Button(frame_select, text="Regresar al Menú", command=regresar, bg="#6a0dad", fg="white")
    regresar_button.grid(row=4, column=3, padx=5, pady=10, sticky="ew")

    # Frame para contener el canvas y la barra de desplazamiento
    scroll_frame = tk.Frame(root)
    scroll_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    # Canvas para contener las tablas
    tables_canvas = tk.Canvas(scroll_frame)
    tables_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Barra de desplazamiento vertical
    scrollbar = tk.Scrollbar(scroll_frame, orient="vertical", command=tables_canvas.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Enlazar el canvas con la barra de desplazamiento
    tables_canvas.config(yscrollcommand=scrollbar.set)

    # Frame para las tablas en el canvas
    tables_canvas_frame = tk.Frame(tables_canvas)
    tables_canvas_frame.bind("<Configure>", lambda e: tables_canvas.configure(scrollregion=tables_canvas.bbox("all")))

    # Crear ventana interna del canvas
    tables_canvas.create_window((0, 0), window=tables_canvas_frame, anchor="nw")

    root.mainloop()