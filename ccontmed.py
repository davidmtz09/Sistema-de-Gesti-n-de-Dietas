import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from calendar import monthrange
import pandas as pd
import webbrowser
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import tempfile
import openpyxl
from openpyxl.styles import PatternFill, Font
import os
import menu_principal
from utilerias import get_file_path
from pathlib import Path

def nombre_a_numero_mes(nombre_mes):
    nombres_meses = {
        "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4, "Mayo": 5, "Junio": 6,
        "Julio": 7, "Agosto": 8, "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12
    }
    return nombres_meses.get(nombre_mes, 0)  # Retorna 0 si el mes no es válido

def regresar():
    root.destroy()
    menu_principal.mostrar_menu()

def contmed():
    global root

    # Variables globales para almacenar los totales del mes
    total_mes = {
        "Internos": {"DESAYUNO": 0, "COMIDA": 0, "CENA": 0},
        "Residentes": {"DESAYUNO": 0, "COMIDA": 0, "CENA": 0}
    }

    def obtener_nombre_mes(mes):
        meses = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]
        return meses[mes - 1]

    # Función para contar alimentos por día
    def contar_alimentos_por_dia(df, filtro_columna_8, mes, anho, dia):
        conteos = {"DESAYUNO": 0, "COMIDA": 0, "CENA": 0}
        df_filtrado = df[df.iloc[:, 8] == filtro_columna_8]

        for _, row in df_filtrado.iterrows():
            periodo = row.iloc[10]
            tipo_alimento = row.iloc[9]

            try:
                fecha_inicio, fecha_fin = periodo.split(' - ')
                fecha_inicio = pd.to_datetime(fecha_inicio, dayfirst=True)
                fecha_fin = pd.to_datetime(fecha_fin, dayfirst=True)

                # Verificar si el rango de fechas incluye el día específico del mes y año seleccionados
                if fecha_inicio <= pd.Timestamp(anho, mes, dia) <= fecha_fin:
                    conteos[tipo_alimento] += 1

            except ValueError:
                continue

        return conteos
    # Función para actualizar las tablas
    def update_tables():
        try:
            year = int(year_var.get())
            month = nombre_a_numero_mes(month_var.get())
            num_days = monthrange(year, month)[1]

            # Limpiar el contenido del canvas
            for widget in tables_canvas_frame.winfo_children():
                widget.destroy()

            # Cargar el archivo de Excel
            medicos_df = pd.read_excel(Path.home() / "ContPerFactura" / "medicos.xlsx")

            # Inicializar conteos totales para todo el mes
            total_mes = {
                "Internos": {"DESAYUNO": 0, "COMIDA": 0, "CENA": 0},
                "Residentes": {"DESAYUNO": 0, "COMIDA": 0, "CENA": 0}
            }

            # Crear las tablas dinámicas para cada día
            for day in range(1, num_days + 1):
                date_str = f"{day:02d}/{month:02d}/{year}"
                day_frame = tk.Frame(tables_canvas_frame, borderwidth=2, relief="solid", padx=10, pady=10)
                day_frame.grid(row=(day-1)//3, column=(day-1)%3, padx=5, pady=5, sticky="nsew")

                # Crear la tabla para el día
                columns = ["Fecha", "Desayuno", "Comida", "Cena"]
                table = ttk.Treeview(day_frame, columns=columns, show="headings", height=4)

                # Configurar encabezados de la tabla
                table.heading("Fecha", text=date_str)  # Encabezado de la primera columna con la fecha
                table.column("Fecha", width=150, anchor="center")

                for col in columns[1:]:
                    table.heading(col, text=col)
                    table.column(col, width=100, anchor="center")  # Ajustar el ancho de las columnas

                # Conteos para Internos y Residentes
                internos_conteo = contar_alimentos_por_dia(medicos_df, "MIP", month, year, day)
                residentes_conteo = contar_alimentos_por_dia(medicos_df, "MR", month, year, day)

                # Insertar filas con los conteos
                table.insert('', 'end', values=("Internos", internos_conteo["DESAYUNO"], internos_conteo["COMIDA"], internos_conteo["CENA"]))
                table.insert('', 'end', values=("Residentes", residentes_conteo["DESAYUNO"], residentes_conteo["COMIDA"], residentes_conteo["CENA"]))

                # Calcular subtotales y total
                subtotal_desayuno = internos_conteo["DESAYUNO"] + residentes_conteo["DESAYUNO"]
                subtotal_comida = internos_conteo["COMIDA"] + residentes_conteo["COMIDA"]
                subtotal_cena = internos_conteo["CENA"] + residentes_conteo["CENA"]
                total_consumos = subtotal_desayuno + subtotal_comida + subtotal_cena

                table.insert('', 'end', values=("Subtotal", subtotal_desayuno, subtotal_comida, subtotal_cena),tags=('subtotales',))
                table.insert('', 'end', values=("TOTAL", "", total_consumos, ""),tags=('total_consumos',))
                table.tag_configure('subtotales', background="#c965d3", foreground="white")
                table.tag_configure('total_consumos', background="purple", foreground="white")

                table.pack(padx=5, pady=5, fill=tk.BOTH, expand=True)

                # Ajustar el tamaño del marco del día para que la tabla ocupe todo el ancho disponible
                day_frame.update_idletasks()
                day_frame_width = table.winfo_width()
                day_frame.config(width=day_frame_width)

                # Sumar a los totales del mes
                total_mes["Internos"]["DESAYUNO"] += internos_conteo["DESAYUNO"]
                total_mes["Internos"]["COMIDA"] += internos_conteo["COMIDA"]
                total_mes["Internos"]["CENA"] += internos_conteo["CENA"]
                total_mes["Residentes"]["DESAYUNO"] += residentes_conteo["DESAYUNO"]
                total_mes["Residentes"]["COMIDA"] += residentes_conteo["COMIDA"]
                total_mes["Residentes"]["CENA"] += residentes_conteo["CENA"]

            # Crear la tabla fija al final con el total del mes
            fixed_table_frame = tk.Frame(tables_canvas_frame, borderwidth=2, relief="solid", padx=10, pady=10)
            fixed_table_frame.grid(row=(num_days)//3 + 1, column=0, padx=5, pady=5, sticky="nsew")

            fixed_columns = ["Total Gral.", "Desayuno", "Comida", "Cena"]
            fixed_table = ttk.Treeview(fixed_table_frame, columns=fixed_columns, show="headings", height=4)

            # Configurar encabezados de la tabla fija
            fixed_table.heading("Total Gral.", text="Total Gral.")
            fixed_table.column("Total Gral.", width=150, anchor="center")

            for col in fixed_columns[1:]:
                fixed_table.heading(col, text=col)
                fixed_table.column(col, width=100, anchor="center")

            # Insertar filas con los totales del mes
            fixed_table.insert('', 'end', values=("Internos", total_mes["Internos"]["DESAYUNO"], total_mes["Internos"]["COMIDA"], total_mes["Internos"]["CENA"]))
            fixed_table.insert('', 'end', values=("Residentes", total_mes["Residentes"]["DESAYUNO"], total_mes["Residentes"]["COMIDA"], total_mes["Residentes"]["CENA"]))

            # Calcular subtotales y total del mes
            subtotal_desayuno_mes = total_mes["Internos"]["DESAYUNO"] + total_mes["Residentes"]["DESAYUNO"]
            subtotal_comida_mes = total_mes["Internos"]["COMIDA"] + total_mes["Residentes"]["COMIDA"]
            subtotal_cena_mes = total_mes["Internos"]["CENA"] + total_mes["Residentes"]["CENA"]
            total_consumos_mes = subtotal_desayuno_mes + subtotal_comida_mes + subtotal_cena_mes
        
            fixed_table.insert('', 'end', values=("Subtotal", subtotal_desayuno_mes, subtotal_comida_mes, subtotal_cena_mes),tags=('subtotales',))
            # Asignar el tag al estilo que creaste
            fixed_table.insert('', 'end', values=("TOTAL", "", total_consumos_mes, ""),tags=('total_consumos',))
            fixed_table.pack(padx=5, pady=5, fill=tk.BOTH, expand=True)

            fixed_table.tag_configure('subtotales', background="#c965d3", foreground="white")
            fixed_table.tag_configure('total_consumos', background="purple", foreground="white")

            # Ajustar el tamaño del marco de la tabla fija para que la tabla ocupe todo el ancho disponible
            fixed_table_frame.update_idletasks()
            fixed_table_frame_width = fixed_table.winfo_width()
            fixed_table_frame.config(width=fixed_table_frame_width)

            # Actualizar el tamaño del canvas
            tables_canvas.configure(scrollregion=tables_canvas.bbox("all"))

        except ValueError:
            messagebox.showerror("Error", "Asegúrese de que el mes y el año estén seleccionados correctamente.")
        except FileNotFoundError as e:
            messagebox.showerror("Error", f"No se pudo encontrar el archivo: {str(e)}")

    def exportar_pdf():
        try:
            mes = nombre_a_numero_mes(month_var.get())
            anho = int(year_var.get())
            nombre_mes = obtener_nombre_mes(mes)
            
            # Nombre del archivo PDF
            file_name = get_file_path("reporte.pdf")
            pdf = SimpleDocTemplate(file_name, pagesize=letter)
            
            elements = []
            styles = getSampleStyleSheet()

            # Encabezado del hospital
            header_text = """
            HOSPITAL GENERAL DE CUERNAVACA DR. JOSE G. PARRES<br/>
            SUBDIRECCIÓN MÉDICA<br/>
            DEPARTAMENTO DE NUTRICIÓN
            """
            # Aplicar estilo normal (céntrico)
            header_paragraph = Paragraph(f"<para align='center'>{header_text}</para>", styles['Normal'])
            elements.append(header_paragraph)
            elements.append(Paragraph("<br/><br/>", styles['Normal']))
            
            # Título
            title = Paragraph(f"Concentrado Mensual de Alimentos<br/> Medicos Internos y Residentes<br/>{nombre_mes}-{anho}", styles['Title'])
            elements.append(title)
            elements.append(Paragraph("<br/>", styles['Normal']))

            # Recorrer todas las tablas en la interfaz y agregar al PDF
            widgets = tables_canvas_frame.winfo_children()
            num_tables = len(widgets)
            
            for idx, widget in enumerate(widgets, start=1):
                table_data = []
                
                # Obtener los datos de la tabla en la interfaz
                tree = widget.winfo_children()[0]  # El Treeview es el primer widget en el frame
                col_names = tree["columns"]
                
                # Modificar encabezado para mostrar "Día X" o "Total Gral:" según corresponda
                if idx == num_tables:
                    table_data.append(["Total Gral:"] + list(col_names[1:]))
                else:
                    table_data.append([f"Día {idx}"] + list(col_names[1:]))
                
                for child in tree.get_children():
                    row = tree.item(child)['values']
                    table_data.append(list(row))

                # Crear tabla para el PDF
                table = Table(table_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch])
                
                # Agregar estilo a la tabla
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
                    ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ]))
                
                # Añadir la tabla a los elementos
                elements.append(table)
                elements.append(Paragraph("<br/>", styles['Normal']))
            
            # Generar el PDF
            pdf.build(elements)
            
            # Abrir el PDF generado
            webbrowser.open(file_name)
            
        except ValueError as ve:
            print(f"Error de valor: {ve}")  # Para depurar errores de conversión
            messagebox.showerror("Error", "Asegúrese de que el mes y el año estén seleccionados correctamente.")
        except FileNotFoundError as e:
            messagebox.showerror("Error", f"No se pudo encontrar el archivo: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"Ha ocurrido un error: {str(e)}")

    def exportar_excel():
        try:
            # Crear un archivo temporal
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            temp_file.close()

            # Crear un nuevo libro de trabajo
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte MRMIP"

            # Definir estilos para las celdas
            header_fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
            total_fill = PatternFill(start_color="FCF75E", end_color="FCF75E", fill_type="solid")
            total_font = Font(color="1B101D")
            header_font = Font(color="FFFFFF")
            
            # Obtener los widgets (tablas) de la interfaz
            widgets = tables_canvas_frame.winfo_children()

            row_num = 1
            num_tables = len(widgets)

            for idx, widget in enumerate(widgets, start=1):
                tree = widget.winfo_children()[0]  # El Treeview es el primer widget en el frame

                # Modificar el primer encabezado según sea "Día X" o "Total Gral"
                if idx == num_tables:
                    col_names = ["Total Gral"] + list(tree["columns"][1:])
                else:
                    col_names = [f"Día {idx}"] + list(tree["columns"][1:])

                # Insertar encabezados
                for col_num, col_name in enumerate(col_names, start=1):
                    cell = ws.cell(row=row_num, column=col_num, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font

                row_num += 1

                # Insertar datos de las filas
                for child in tree.get_children():
                    row = tree.item(child)['values']
                    for col_num, value in enumerate(row, start=1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        # Aplicar estilo específico para la fila 'TOTAL'
                        if row[0] == "TOTAL":
                            cell.fill = total_fill
                            cell.font = total_font
                    row_num += 1

                row_num += 1  # Añadir una fila vacía entre tablas

            # Ajustar el ancho de las columnas
            for col in ws.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                adjusted_width = (max_length + 2)
                ws.column_dimensions[col[0].column_letter].width = adjusted_width

            # Guardar el archivo y abrirlo automáticamente
            wb.save(temp_file.name)
            webbrowser.open(temp_file.name)

        except Exception as e:
            messagebox.showerror("Error", f"Ha ocurrido un error: {str(e)}")

    # Crear ventana
    root = tk.Tk()
    root.title("CONCENTRADO DE ALIMENTOS DE MEDICOS INTERNOS Y RESIDENTES")
    root.geometry("1200x600")

    # Agregar estilos al Treeview
    style = ttk.Style()
    # Tema
    style.theme_use("clam")
    # Configuración de los colores del treeview
    style.configure("Treeview",
        background="#b7eefa",
        foreground="black",
        fieldbackground="#b7eefa"
    )

    style.map("Treeview",
        background=[('selected', '#bd1323')],
        foreground=[('selected', 'white')]
    )

    # Variables de control
    month_var = tk.StringVar()
    year_var = tk.StringVar()

    # Frame para la selección de mes y año
    frame_select = tk.Frame(root)
    frame_select.pack(pady=10)

     # Etiqueta principal en el centro
    tk.Label(frame_select, text="HOSPITAL GENERAL DE CUERNAVACA DR. JOSE G. PARRES \n SUBDIRECCIÓN MÈDICA \n DEPARTAMENTO DE NUTRICIÓN", 
            font=('Arial', 10), anchor="center", justify="center", bg=frame_select.cget("background")).grid(row=0, column=0, columnspan=4, padx=5, pady=10, sticky="ew")

    # Etiqueta del título
    tk.Label(frame_select, text="CONCENTRADO MENSUAL DE ALIMENTOS DE MEDICOS INTERNOS Y RESIDENTES", 
            font=('Verdana', 12, "bold"), anchor="center", justify="center", bg=frame_select.cget("background")).grid(row=1, column=0, columnspan=4, padx=5, pady=10, sticky="ew")

    # Etiqueta y menú desplegable para el mes
    tk.Label(frame_select, text="Seleccione Mes:").grid(row=2, column=1, padx=5, pady=5, sticky="e")
    month_menu = ttk.Combobox(frame_select, textvariable=month_var, values=[
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ], state="readonly")
    month_menu.grid(row=2, column=2, padx=5, pady=5, sticky="w")

    # Etiqueta y menú desplegable para el año
    tk.Label(frame_select, text="Seleccione Año:").grid(row=3, column=1, padx=5, pady=5, sticky="e")
    year_menu = ttk.Combobox(frame_select, textvariable=year_var, values=[str(year) for year in range(2024, 2051)], state="readonly")
    year_menu.grid(row=3, column=2, padx=5, pady=5, sticky="w")

    # Botón para actualizar las tablas
    update_button = tk.Button(frame_select, text="Actualizar", command=update_tables, bg="#6a0dad", fg="white")
    update_button.grid(row=4, column=0, padx=5, pady=10, sticky="ew")

    #Botón para exportar a pdf
    update_button = tk.Button(frame_select, text="Exportar a PDF", command=exportar_pdf, bg="#bd1323", fg="white")
    update_button.grid(row=4, column=1, padx=5, pady=10, sticky="ew")

    #Botón para exportar a excel
    update_button = tk.Button(frame_select, text="Exportar a EXCEL", command=exportar_excel, bg="green", fg="white")
    update_button.grid(row=4, column=2, padx=5, pady=10, sticky="ew")

    regresar_button = tk.Button(frame_select, text="Regresar al Menú", command=regresar, bg="#6a0dad", fg="white")
    regresar_button.grid(row=4, column=3, padx=5, pady=10, sticky="ew")

    # Frame para contener el canvas y la barra de desplazamiento
    scroll_frame = tk.Frame(root)
    scroll_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    # Canvas para contener las tablas
    tables_canvas = tk.Canvas(scroll_frame)
    tables_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Barra de desplazamiento vertical
    scrollbar = tk.Scrollbar(scroll_frame, orient="vertical", command=tables_canvas.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Enlazar el canvas con la barra de desplazamiento
    tables_canvas.config(yscrollcommand=scrollbar.set)

    # Frame para las tablas en el canvas
    tables_canvas_frame = tk.Frame(tables_canvas)
    tables_canvas_frame.bind("<Configure>", lambda e: tables_canvas.configure(scrollregion=tables_canvas.bbox("all")))

    # Crear ventana interna del canvas
    tables_canvas.create_window((0, 0), window=tables_canvas_frame, anchor="nw")

    root.mainloop()